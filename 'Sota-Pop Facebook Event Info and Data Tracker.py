import time
import datetime
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import os
import time

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

wb = Workbook()

dateandtime = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
dest_filename = str(dateandtime) + ".xlsx"
print(dateandtime)

def geteventlist(facebookpage):
    # url = input("enter url of page who's eventclicks you would like to scrub")
    assert isinstance(facebookpage, object)
    driver.get(facebookpage)
    time.sleep(5)  # Let the user actually see something!
    print("k")
    scroll()
    clicknotnow = driver.find_element_by_xpath('//*[@id="expanding_cta_close_button"]')
    clicknotnow.click()
    scroll()
    driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + Keys.HOME)
    scroll()
    elems = driver.find_elements_by_class_name("_4dmk")
    #eventnums = []
    for elem in elems:
        h2_element = elem
        a_element = h2_element.find_element_by_tag_name("a")
        href = a_element.get_attribute('href')
        hrefstr = str(href)
        hrefstr = hrefstr.replace("https://www.facebook.com/events/", "")
        sep = '/'
        eventnum = hrefstr.split(sep, 1)[0]
        print(eventnum)
        eventnums.append(eventnum)
    print(eventnums)


ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

savelocationstring = "C:\\Python\\'SOTA-POP EVENT DATA BY DATE PULLED\\"


wb.save(savelocationstring + dest_filename)